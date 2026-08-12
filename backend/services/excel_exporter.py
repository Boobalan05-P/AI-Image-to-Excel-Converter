import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)

def convert_matrix_to_dataframe(table_matrix: list) -> pd.DataFrame:
    """
    Converts 2D list of strings or dicts into a cleaned pandas DataFrame.
    Automatically assigns column headers if present or generates generic headers.
    """
    if not table_matrix:
        return pd.DataFrame()

    raw_data = []
    for row in table_matrix:
        row_values = []
        for cell in row:
            if isinstance(cell, dict):
                val = cell.get("text", "")
            else:
                val = str(cell)
            row_values.append(val.strip())
        raw_data.append(row_values)

    # Normalize row lengths to match max column count
    max_cols = max(len(r) for r in raw_data) if raw_data else 0
    if max_cols == 0:
        return pd.DataFrame()

    for r in raw_data:
        while len(r) < max_cols:
            r.append("")

    # Images do not reliably tell us whether the first line is a heading. Never
    # silently consume it: every extracted row must appear in the export.
    headers = [f"Column {i+1}" for i in range(max_cols)]
    df_data = raw_data

    df = pd.DataFrame(df_data, columns=headers)

    # Keep OCR values as text. Long numbers are often IDs or phone numbers, and
    # converting them to numbers can remove leading zeroes or change Excel's
    # display (for example, to scientific notation).

    return df

def generate_styled_excel(df: pd.DataFrame, output_path: str, sheet_name: str = "Sheet1", include_header: bool = False):
    """
    Generates a beautifully styled Excel workbook (.xlsx) using OpenPyXL.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    # Write dataframe to sheet
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=include_header), 1):
        ws.append(r)
        is_header = include_header and r_idx == 1

        for c_idx in range(1, len(r) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border

            if is_header:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            else:
                cell.font = cell_font
                if r_idx % 2 == 0:
                    cell.fill = zebra_fill
                
                # Align numbers right, text left
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    wb.save(output_path)
    logger.info(f"Excel successfully saved to {output_path}")

def generate_csv(df: pd.DataFrame, output_path: str, include_header: bool = False):
    """Saves DataFrame as CSV file."""
    df.to_csv(output_path, index=False, header=include_header)
    logger.info(f"CSV successfully saved to {output_path}")

def merge_tables_to_excel(tables_list: list, output_path: str, mode: str = "sheets"):
    """
    Merges multiple table DataFrames into a single Excel workbook.
    mode: 'sheets' (separate tab per image) or 'merged' (single combined sheet).
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    if mode == "sheets":
        for idx, item in enumerate(tables_list):
            title = item.get("filename", f"Table_{idx+1}")[:30] # Excel sheet name limit is 31 chars
            title = "".join(c for c in title if c not in r"[]:*?/\\")
            df = convert_matrix_to_dataframe(item.get("table_data", []))
            ws = wb.create_sheet(title=title or f"Sheet_{idx+1}")
            
            # Format sheet
            for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                ws.append(r)
                
            # Auto-adjust columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            ws.views.sheetView[0].showGridLines = True
    else:
        # Stacked into single combined sheet
        ws = wb.create_sheet(title="Combined_Tables")
        current_row = 1
        for idx, item in enumerate(tables_list):
            df = convert_matrix_to_dataframe(item.get("table_data", []))
            if not df.empty:
                # Add table title banner
                ws.cell(row=current_row, column=1, value=f"Source: {item.get('filename', f'Table {idx+1}')}")
                ws.cell(row=current_row, column=1).font = Font(size=12, bold=True, color="1E40AF")
                current_row += 1
                
                for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), current_row):
                    ws.append(r)
                current_row += len(df) + 3 # Space between tables

        ws.views.sheetView[0].showGridLines = True

    wb.save(output_path)
    logger.info(f"Merged Excel file saved to {output_path}")
